#!/usr/bin/env python3
"""Migrate Excel data to SQLite database."""

import sqlite3
import openpyxl
from datetime import datetime
import os
import sys

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else '/root/.claude/uploads/53eb8ac1-c318-422e-ab70-338ecb0f4aab/532c3ff2-BLOCKS_SALES.xlsx'
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'sales.db')

os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

cur.execute('''
CREATE TABLE IF NOT EXISTS sales (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  date TEXT NOT NULL,
  customer_name TEXT,
  address TEXT,
  phone TEXT,
  size INTEGER NOT NULL,
  quantity INTEGER NOT NULL,
  rate REAL,
  amount REAL,
  advance REAL DEFAULT 0,
  balance REAL DEFAULT 0,
  status TEXT DEFAULT 'CLOSED',
  payment_mode TEXT,
  notes TEXT,
  month_label TEXT,
  created_at TEXT DEFAULT (datetime('now'))
)
''')

cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_date ON sales(date)')
cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_status ON sales(status)')
cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_customer ON sales(customer_name)')
cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_month ON sales(month_label)')

wb = openpyxl.load_workbook(EXCEL_PATH)

# Old format sheets (Jul-Nov 2025): columns = SL,NO | DATE | PARTY | 4'' | 6'' | 8'' | CASH | NIYAKRISH | MKL | KMK
OLD_SHEETS = ['JULY-25', 'AUG-25', 'SEPT-25', 'OCTO-25', 'NOV-2025']
# New format sheets (Dec+): S.NO | DATE | CUSTOMER | ADDRESS | CELL | SIZE | QTY | RATE | AMOUNT | ADVANCE | BALANCE | STATUS | MODE
NEW_SHEETS = ['DEC', 'JAN-2026', 'FEB-26', 'MAR-26', 'Apr-26', 'MAY-26']

def safe_float(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v.startswith('=') or v == '':
            return None
        try:
            return float(v.replace(',', ''))
        except:
            return None
    try:
        return float(v)
    except:
        return None

def safe_int(v):
    f = safe_float(v)
    return int(f) if f is not None else None

def fmt_date(d):
    if isinstance(d, datetime):
        return d.strftime('%Y-%m-%d')
    if isinstance(d, str) and d:
        return d
    return None

def payment_mode_old(row, start_idx=6):
    # columns: CASH, NIYAKRISH AC, MKL AC, KMK AC
    modes = ['CASH', 'NY A/C', 'MKL A/C', 'KMK A/C']
    for i, mode in enumerate(modes):
        val = row[start_idx + i] if start_idx + i < len(row) else None
        if val is not None:
            return mode
    return None

total = 0

for sheet_name in OLD_SHEETS:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    for row in rows[2:]:  # skip title + header
        if not row or row[0] is None:
            continue
        date_val = fmt_date(row[1])
        if not date_val:
            continue
        party = str(row[2]).strip() if row[2] else None
        qty_4 = safe_int(row[3])
        qty_6 = safe_int(row[4])
        qty_8 = safe_int(row[5])

        # Determine payment across columns
        cash = safe_float(row[6]) if len(row) > 6 else None
        ny_ac = safe_float(row[7]) if len(row) > 7 else None
        mkl_ac = safe_float(row[8]) if len(row) > 8 else None
        kmk_ac = safe_float(row[9]) if len(row) > 9 else None

        def get_payment_mode_and_amount():
            if cash:
                return 'CASH', cash
            if ny_ac:
                return 'NY A/C', ny_ac
            if mkl_ac:
                return 'MKL A/C', mkl_ac
            if kmk_ac:
                return 'KMK A/C', kmk_ac
            return None, None

        mode, amount = get_payment_mode_and_amount()

        # Insert one row per block size
        entries = []
        if qty_4:
            entries.append((4, qty_4))
        if qty_6:
            entries.append((6, qty_6))
        if qty_8:
            entries.append((8, qty_8))

        # If multiple sizes, split amount proportionally (rough)
        total_qty = sum(q for _, q in entries)
        for size, qty in entries:
            amt = None
            if amount and total_qty:
                amt = round(amount * qty / total_qty, 2)
            cur.execute('''
                INSERT INTO sales (date, customer_name, size, quantity, amount, payment_mode, status, month_label)
                VALUES (?, ?, ?, ?, ?, ?, 'CLOSED', ?)
            ''', (date_val, party, size, qty, amt, mode, sheet_name))
            count += 1
    print(f'{sheet_name}: {count} entries')
    total += count

for sheet_name in NEW_SHEETS:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        date_val = fmt_date(row[1])
        if not date_val:
            continue
        customer = str(row[2]).strip() if row[2] else None
        address = str(row[3]).strip() if row[3] else None
        phone = str(int(row[4])) if row[4] and isinstance(row[4], (int, float)) else (str(row[4]).strip() if row[4] else None)
        size = safe_int(row[5])
        qty = safe_int(row[6])
        rate = safe_float(row[7])
        amount = safe_float(row[8])
        advance = safe_float(row[9]) or 0
        balance = safe_float(row[10]) or 0
        status = str(row[11]).strip().upper() if row[11] else 'CLOSED'
        mode = str(row[12]).strip() if row[12] else None
        notes = str(row[13]).strip() if len(row) > 13 and row[13] else None

        if size is None or qty is None:
            continue

        # Compute amount from rate*qty if missing
        if amount is None and rate and qty:
            amount = round(rate * qty, 2)

        # Compute balance if missing
        if balance == 0 and amount and advance is not None:
            balance = max(0, amount - advance)

        cur.execute('''
            INSERT INTO sales (date, customer_name, address, phone, size, quantity, rate, amount,
                               advance, balance, status, payment_mode, notes, month_label)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (date_val, customer, address, phone, size, qty, rate, amount,
              advance, balance, status, mode, notes, sheet_name))
        count += 1
    print(f'{sheet_name}: {count} entries')
    total += count

conn.commit()
conn.close()
print(f'\nTotal migrated: {total} entries')
